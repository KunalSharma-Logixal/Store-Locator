import os
import json
import uuid
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from typing import List, Dict, Any, Tuple, Generator, Optional
from backend.utils.import_specs.base_spec import EntityImportSpec
from backend.utils.header_transformer import HeaderTransformer

class GenericImporter:
    def __init__(self, spec: EntityImportSpec, chunk_size: int = 50, reports_dir: str = None, temp_base_dir: str = None):
        self.spec = spec
        self.chunk_size = chunk_size
        self.reports_dir = reports_dir
        self.temp_base_dir = temp_base_dir or reports_dir
        self.transformer = HeaderTransformer(spec.header_map)

    def import_file(self, file_content: bytes, filename: str) -> Dict[str, Any]:
        # Extension checking
        ext = os.path.splitext(filename)[1].lower()
        if ext not in [".csv", ".xlsx"]:
            return {
                "success": False,
                "message": f"Unsupported file type '{ext}'. Only CSV and XLSX are allowed.",
                "summary": self._empty_summary("00000000000000000000000000000000", "Failed")
            }

        bulk_no = uuid.uuid4().hex
        os.makedirs(self.reports_dir, exist_ok=True)
        temp_filepath = os.path.join(self.temp_base_dir, f"temp_{bulk_no}{ext}")

        try:
            with open(temp_filepath, "wb") as f:
                f.write(file_content)
        except Exception as e:
            return {
                "success": False,
                "message": f"Failed to cache file on server: {str(e)}",
                "summary": self._empty_summary(bulk_no, "Failed")
            }

        try:
            if ext == ".csv":
                headers, row_gen = self._stream_csv_rows(temp_filepath)
            else:
                headers, row_gen = self._stream_excel_rows(temp_filepath)
        except Exception as e:
            self._cleanup_temp_file(temp_filepath)
            return {
                "success": False,
                "message": f"File parsing failed: {str(e)}",
                "summary": self._empty_summary(bulk_no, "Failed")
            }

        # Transform headers
        normalized_cols = self.transformer.transform(headers)

        # Check missing required columns
        missing_fields = [f for f in self.spec.required_fields if normalized_cols.get(f) is None]

        if missing_fields:
            self._cleanup_temp_file(temp_filepath)
            return {
                "success": False,
                "message": "Header validation failed.",
                "summary": {
                    "bulk_no": bulk_no,
                    "total_rows": 0,
                    "successful_rows": 0,
                    "failed_rows": 0,
                    "duplicate_rows": 0,
                    "date": datetime.now().isoformat(),
                    "processing_status": "Failed",
                    "errors": [{"row": 0, "error": f"Missing required columns: {', '.join(missing_fields)}"}]
                }
            }

        # Initialize duplicates tracker from database
        seen_ids, seen_combos = self.spec.initialize_duplicates_tracker()

        # Stats trackers
        stats = {
            "total_rows": 0,
            "successful_rows": 0,
            "failed_rows": 0,
            "duplicate_rows": 0
        }

        failed_records_list = []
        chunk_buffer = []

        try:
            for row_idx, row_dict in enumerate(row_gen, start=1):
                chunk_buffer.append((row_idx, row_dict))
                if len(chunk_buffer) >= self.chunk_size:
                    self._process_chunk(chunk_buffer, normalized_cols, seen_ids, seen_combos, failed_records_list, stats)
                    chunk_buffer = []
            
            # Process remaining
            if chunk_buffer:
                self._process_chunk(chunk_buffer, normalized_cols, seen_ids, seen_combos, failed_records_list, stats)
        except Exception as e:
            self._cleanup_temp_file(temp_filepath)
            return {
                "success": False,
                "message": f"Processing interrupted: {str(e)}",
                "summary": self._empty_summary(bulk_no, "Failed")
            }

        self._cleanup_temp_file(temp_filepath)

        # Generate failed records CSV report
        failed_file_name = None
        if failed_records_list:
            failed_file_name = f"failed_rows_{bulk_no}.csv"
            failed_csv_path = os.path.join(self.reports_dir, failed_file_name)
            try:
                error_cols = [c for c in failed_records_list[0].keys() if c != "Error"] + ["Error"]
                df_failed = pd.DataFrame(failed_records_list)
                df_failed = df_failed[error_cols]
                df_failed.to_csv(failed_csv_path, index=False, encoding="utf-8")
            except Exception as e:
                print(f"Error saving failed records CSV: {str(e)}")

        status_str = "Completed" if stats["failed_rows"] == 0 else "Partially Failed"
        if stats["successful_rows"] == 0 and stats["total_rows"] > 0:
            status_str = "Failed"

        summary = {
            "bulk_no": bulk_no,
            "total_rows": stats["total_rows"],
            "successful_rows": stats["successful_rows"],
            "failed_rows": stats["failed_rows"],
            "duplicate_rows": stats["duplicate_rows"],
            "date": datetime.now().isoformat(),
            "processing_status": status_str,
            "failed_rows_file": failed_file_name
        }

        # Save summary JSON file
        summary_path = os.path.join(self.reports_dir, f"summary_{bulk_no}.json")
        try:
            with open(summary_path, "w", encoding="utf-8") as sf:
                json.dump(summary, sf, indent=2, ensure_ascii=False)
        except Exception as e:
            print(f"Error saving summary JSON file: {str(e)}")

        return {
            "success": True,
            "message": "Upload processing completed.",
            "summary": summary
        }

    def _process_chunk(
        self,
        chunk: List[Tuple[int, Dict[str, Any]]],
        normalized_cols: Dict[str, Any],
        seen_ids: set,
        seen_combos: set,
        failed_records_list: List[Dict[str, Any]],
        stats: Dict[str, int]
    ) -> None:
        valid_records = []

        for row_num, row_dict in chunk:
            stats["total_rows"] += 1
            
            # Validate row-level rules
            row_errors = self.spec.validate_row(row_num, row_dict, normalized_cols)

            if row_errors:
                stats["failed_rows"] += 1
                failed_rec = dict(row_dict)
                failed_rec["Error"] = " | ".join(row_errors)
                failed_records_list.append(failed_rec)
                continue

            # Check duplicates
            is_dup, resolved_id, combo_key = self.spec.check_duplicate(row_dict, normalized_cols, seen_ids, seen_combos)

            if is_dup:
                stats["failed_rows"] += 1
                stats["duplicate_rows"] += 1
                failed_rec = dict(row_dict)
                failed_rec["Error"] = f"Duplicate {self.spec.entity_name} detected."
                failed_records_list.append(failed_rec)
                continue

            # Track seen IDs/combinations
            if resolved_id:
                seen_ids.add(resolved_id)
            if combo_key:
                seen_combos.add(combo_key)

            # Transform raw columns to DB schema
            db_record = self.spec.normalize_row_for_db(row_dict, normalized_cols)
            valid_records.append(db_record)

        # Transact chunk commits
        if valid_records:
            try:
                self.spec.commit_chunk(valid_records)
                stats["successful_rows"] += len(valid_records)
            except Exception as e:
                # If database write fails, record as failed
                for row_num, row_dict in chunk:
                    if self.spec.normalize_row_for_db(row_dict, normalized_cols) in valid_records:
                        stats["failed_rows"] += 1
                        failed_rec = dict(row_dict)
                        failed_rec["Error"] = f"Database commit failure: {str(e)}"
                        failed_records_list.append(failed_rec)

    def _stream_csv_rows(self, filepath: str) -> Tuple[List[str], Generator[Dict[str, Any], None, None]]:
        df_headers = pd.read_csv(filepath, nrows=0)
        headers = [str(col).strip() for col in df_headers.columns]

        def row_generator():
            chunks = pd.read_csv(filepath, chunksize=100)
            for chunk in chunks:
                chunk = chunk.where(pd.notnull(chunk), None)
                for _, row in chunk.iterrows():
                    yield row.to_dict()

        return headers, row_generator()

    def _stream_excel_rows(self, filepath: str) -> Tuple[List[str], Generator[Dict[str, Any], None, None]]:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        sheet = wb.active
        row_iter = sheet.iter_rows(values_only=True)

        try:
            raw_headers = next(row_iter)
        except StopIteration:
            wb.close()
            raise ValueError("Spreadsheet has no rows.")

        if not raw_headers:
            wb.close()
            raise ValueError("Spreadsheet header row is empty.")

        headers = [str(h).strip() if h is not None else "" for h in raw_headers]

        def row_generator():
            try:
                for row in row_iter:
                    if not row or not any(cell is not None for cell in row):
                        continue
                    row_dict = {}
                    for idx, val in enumerate(row):
                        if idx < len(headers):
                            row_dict[headers[idx]] = val
                    yield row_dict
            finally:
                wb.close()

        return headers, row_generator()

    def _cleanup_temp_file(self, path: str) -> None:
        if os.path.exists(path):
            try:
                os.remove(path)
            except OSError:
                pass

    def _empty_summary(self, bulk_no: str, status: str) -> Dict[str, Any]:
        return {
            "bulk_no": bulk_no,
            "total_rows": 0,
            "successful_rows": 0,
            "failed_rows": 0,
            "duplicate_rows": 0,
            "date": datetime.now().isoformat(),
            "processing_status": status,
            "failed_rows_file": None
        }
